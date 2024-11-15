from django.shortcuts import render
from django.http import HttpResponse
import pandas as pd
from .forms import TeamForm
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

def generate_excel(request):
    if request.method == 'POST':
        form = TeamForm(request.POST)
        if form.is_valid():
            url = form.cleaned_data['team_name']

            df_schedule, li_MMP, li_FMP = scrape_page(url)
            wb = make_sheet(df_schedule, li_MMP, li_FMP)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=team.xlsx'
            wb.save(response)
            return response
    else:
        form = TeamForm()
    
    return render(request, 'myapp/team_form.html', {'form': form})


# Function to adjust the column width to fit the longest text
def adjust_all_columns_width(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        # Iterate over all cells in this column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        # Set the column width based on the maximum length found
        sheet.column_dimensions[col_letter].width = max_length + 2

def make_sheet(df_schedule, li_MMP, li_FMP):
    df_schedule["Time"] = df_schedule["When"].apply(lambda x : x.split("\r\n")[1].strip())
    df_schedule["Date"] = df_schedule["When"].apply(lambda x : x.split("\r\n")[0].strip())
    df_schedule["Opponent"] = df_schedule["Opponent"].apply(lambda x: x.split("(")[0].strip())
    df_schedule["Field"] = df_schedule["Field"].apply(lambda x : x.split("\r\n")[0].strip())

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    start_row = 1

    for headerColumn in ["Date", "Time", "Field", "Jersey", "Opponent"]:
        # Define the column headers starting with "Player"
        headers = [headerColumn] + df_schedule[headerColumn].tolist()

        # Write the headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col_num, value=header)

        start_row += 1


    mixed = False
    if len(li_MMP) > 0 and len(li_FMP) > 0:
        mixed = True

    # Add rows for Total, MMP Total, and FMP Total
    total_row = start_row
    start_row = start_row + 1

    if mixed:
        mmp_total_row = total_row + 1
        fmp_total_row = mmp_total_row + 1
        start_row = start_row + 2


    ws.cell(row=total_row, column=1, value="Total")
    if mixed:
        ws.cell(row=mmp_total_row, column=1, value="MMP Total")
        ws.cell(row=fmp_total_row, column=1, value="FMP Total")
        start_row = start_row + 1

    player_start_row = start_row
    if (mixed):
        player_start_row += 1

    player_end_row = player_start_row + len(li_MMP) + len(li_FMP) - 1
    if (mixed):
        player_end_row += 2

    # Add formulas to calculate the totals for each column
    for col in range(2, len(headers) + 1):
        col_letter = chr(64 + col)
        if not(mixed):
            ws.cell(row=total_row, column=col).value = f'=COUNTIF({col_letter}{player_start_row}:{col_letter}{player_end_row},"Yes")'
        if mixed:
            ws.cell(row=total_row, column=col).value = f'=COUNTIF({col_letter}{player_start_row}:{col_letter}{player_end_row},"Yes")'
            ws.cell(row=mmp_total_row, column=col).value = f'=COUNTIF({col_letter}{player_start_row}:{col_letter}{start_row + len(li_MMP)},"Yes")'
            ws.cell(row=fmp_total_row, column=col).value = f'=COUNTIF({col_letter}{player_end_row - len(li_FMP) + 1}:{col_letter}{player_end_row},"Yes")'


    if (mixed):
        ws.cell(row=start_row, column=1, value="MMP")
        start_row += 1


    # Write the li_MMP items to the worksheet
    for row_num, item in enumerate(li_MMP, start_row):
        ws.cell(row=row_num, column=1, value=item)

    # Add a few rows of space between li_MMP and li_FMP items
    if mixed:
        start_row += len(li_MMP) + 1
        ws.cell(row=start_row, column=1, value="FMP")
        start_row = start_row + 1

    # Write the li_FMP items to the worksheet
    for row_num, item in enumerate(li_FMP, start_row):
        ws.cell(row=row_num, column=1, value=item)

    # Define the data validation for the select list
    select_list = DataValidation(type="list", formula1='"Yes,No,Maybe"', showDropDown=True)

    # Apply the data validation to each cell in the table (excluding headers)
    for row in range(player_start_row, player_end_row + 1):
        for col in range(2, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            ws.add_data_validation(select_list)
            select_list.add(cell)

    adjust_all_columns_width(ws)

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Create rules for "Yes", "No", and "Maybe"
    yes_rule = FormulaRule(formula=[f'B{player_start_row}="Yes"'], fill=green_fill)
    no_rule = FormulaRule(formula=[f'B{player_start_row}="No"'], fill=red_fill)
    maybe_rule = FormulaRule(formula=[f'B{player_start_row}="Maybe"'], fill=yellow_fill)

    ws.conditional_formatting.add(f"B{player_start_row}:{get_column_letter(len(headers) + 1)}{player_end_row}", yes_rule)
    ws.conditional_formatting.add(f"B{player_start_row}:{get_column_letter(len(headers) + 1)}{player_end_row}", no_rule)
    ws.conditional_formatting.add(f"B{player_start_row}:{get_column_letter(len(headers) + 1)}{player_end_row}", maybe_rule)

    under_7_rule = FormulaRule(formula=[f'B6<7'], fill=red_fill)
    atleast_7_rule = FormulaRule(formula=[f'B6>=7'], fill=green_fill)
    ws.conditional_formatting.add(f"B6:{get_column_letter(len(headers) + 1)}6", under_7_rule)
    ws.conditional_formatting.add(f"B6:{get_column_letter(len(headers) + 1)}6", atleast_7_rule)

    if mixed:
        ws.freeze_panes = "A9"
    else:
        ws.freeze_panes = "A7"

    return wb



def scrape_page(url):
    # Send a GET request to the URL
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Find the first element with the class "rgMasterTable"
        rg_master_table = soup.find(class_='rgMasterTable')
        
        # Check if the element was found
        if rg_master_table:
            # Extract table rows
            rows = rg_master_table.find_all('tr')
            
            # Extract headers
            headers = [header.text.strip() for header in rows[0].find_all('th')]
            
            # Extract data
            data = []
            for row in rows[1:]:
                cols = row.find_all(['td', 'th'])
                row_data = []
                for col in cols:
                    colspan = int(col.get('colspan', 1))
                    cell_data = col.text.strip()
                    row_data.append(cell_data)
                    # Append empty strings for the remaining columns
                    row_data.extend([''] * (colspan - 1))
                data.append(row_data)
            
            # Create a DataFrame named df_schedule
            df_schedule = pd.DataFrame(data, columns=headers)
        
        # Find the element with the ID cpMain_cpMain_pnlMen
        pnl_men = soup.find(id='cpMain_cpMain_pnlMen')
        
        # Check if the element was found
        if pnl_men:
            # Find the ol element within this div
            ol_element_men = pnl_men.find('ol')
            
            # Check if the ol element was found
            if ol_element_men:
                # Extract text from each li element and put it in a list called li_MMP
                li_MMP = [li.text.strip().split("\r")[0] for li in ol_element_men.find_all('li')]
            else:
                li_MMP = []
        else:
            li_MMP = []
        
        # Find the element with the ID cpMain_cpMain_pnlWomen
        pnl_women = soup.find(id='cpMain_cpMain_pnlWomen')
        
        # Check if the element was found
        if pnl_women:
            # Find the ol element within this div
            ol_element_women = pnl_women.find('ol')
            
            # Check if the ol element was found
            if ol_element_women:
                # Extract text from each li element and put it in a list called li_FMP
                li_FMP = [li.text.strip().split("\r")[0] for li in ol_element_women.find_all('li')]
            else:
                li_FMP = []
        else:
            li_FMP = []
    return df_schedule, li_MMP, li_FMP
